from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import webview
import PySimpleGUI as sg
import multiprocessing


def getNextRow(sheet):
    nextRow = 1
    for x in sheet.rows:
        if x[1].fill.start_color.rgb == '00000000':
            return nextRow
        nextRow += 1
    return -1

def setInfoText(sheet,currentRow):
    return f'''Applicant Number: {currentRow}\nName: {sheet.cell(currentRow,3).value}, Age: {int(sheet.cell(currentRow,4).value)}\n\nDiscord?: {sheet.cell(currentRow,6).value}, Other Socials: {sheet.cell(currentRow,7).value}\n\nLink to Portfolio: {sheet.cell(currentRow,8).value}\nSpecialty?: {sheet.cell(currentRow,9).value}\n\nOther Skills?: {sheet.cell(currentRow,10).value}\n\nPreferred Positions?: {sheet.cell(currentRow,11).value}\n\nAny Questions?: {sheet.cell(currentRow,12).value}'''

def launchLink(url):
    webview.create_window("Portfolio Link",url)
    chrome_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    session_folder = './webview_session_data'
    webview.start(private_mode=False, storage_path=session_folder, user_agent=chrome_agent)

def main():
    wb = load_workbook(filename="artistresponses.xlsx")
    sheet = wb['Sheet1']

    currentRow = getNextRow(sheet)

    print("SP:",currentRow)

    if currentRow != -1:
        layout = [[sg.Multiline(setInfoText(sheet,currentRow),key="text",font=(None,16),size=(70,25),disabled=True,border_width=0,no_scrollbar=True,text_color="white",background_color=sg.theme_background_color())],
                [sg.Push(),sg.Button("Open Clipboard Link",font=(None,16),key="cliplink"),sg.Push()],
                [sg.Push(),sg.Radio("Nah", group_id=1, font=(None,16), key="nahradio"),sg.Radio("Yeah", group_id=1,font=(None,16), key="yeahradio"),sg.Radio("MOST DEF", group_id=1,font=(None,16),key="defradio"),sg.Push()],
                [sg.Push(),sg.Button("Confirm",font=(None,16), size=(6,1)),sg.Push()]]

        window = sg.Window("Excel Reader",layout)

        while True:
            event, values = window.read()

            if event == sg.WINDOW_CLOSED:
                break
            elif event == "Confirm":
                color = ""
                if values["nahradio"]:
                    color = "FF0000"
                elif values["yeahradio"]:
                    color = "01FF00"
                elif values["defradio"]:
                    color = "38761D"
                else:
                    continue
                sheet.cell(currentRow,2).fill = PatternFill(patternType="solid", fgColor=color)
                print(color)

                currentRow = getNextRow(sheet)
                window["text"].update(setInfoText(sheet,currentRow))
                if currentRow == -1:
                    break
            elif event == "cliplink":
                webview_process = multiprocessing.Process(target=launchLink,args=(sg.clipboard_get(),))
                webview_process.start()

    if currentRow == -1:
        print("All rows completed")

    wb.save('artistresponses.xlsx')
    window.close()

if __name__ == "__main__":
    main()