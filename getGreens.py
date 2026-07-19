from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import webview
import PySimpleGUI as sg
import multiprocessing

# green = FF00FF00
# dark green = FF38761D

def setInfoText(sheet,currentRow):
    return f'''Name: {sheet.cell(currentRow,3).value}, Age: {int(sheet.cell(currentRow,4).value)}\n\nDiscord?: {sheet.cell(currentRow,6).value}, Other Socials: {sheet.cell(currentRow,7).value}\n\nLink to Portfolio: {sheet.cell(currentRow,8).value}\nSpecialty?: {sheet.cell(currentRow,9).value}\n\nOther Skills?: {sheet.cell(currentRow,10).value}\n\nPreferred Positions?: {sheet.cell(currentRow,11).value}\n\nAny Questions?: {sheet.cell(currentRow,12).value}'''

def launchLink(url):
    webview.create_window("Portfolio Link",url)
    chrome_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    session_folder = './webview_session_data'
    webview.start(private_mode=False, storage_path=session_folder, user_agent=chrome_agent)

def main():
    wb = load_workbook(filename="artistresponses.xlsx")
    sheet = wb['Sheet1']

    listGreen = []

    num = 1
    for x in sheet.rows:
        if x[1].fill.start_color.rgb == 'FF00FF00' or x[1].fill.start_color.rgb == 'FF38761D':
            listGreen.append(x)
        num += 1

    greenRow = 0
    currRow = listGreen[greenRow][0].row
    layout = [[sg.Multiline(setInfoText(sheet,currRow),key="text",font=(None,16),size=(70,25),disabled=True,border_width=0,no_scrollbar=True,text_color="white",background_color=sg.theme_background_color())],
                [sg.Push(),sg.Button("Open Clipboard Link",font=(None,16),key="cliplink"),sg.Button("Procreate",font=(None,16),key="procreate"),sg.Button("No Procreate",font=(None,16),key="noprocreate"),sg.Push()]]
    
    window = sg.Window("Green Reader",layout)

    
    listProcreate = []
    while True:
        event, values = window.read()

        if event == sg.WINDOW_CLOSED:
            break
        elif event == "procreate":
            listProcreate.append(listGreen[greenRow])
            greenRow += 1
            if greenRow == len(listGreen):
                break
            currRow = listGreen[greenRow][0].row
            window["text"].update(setInfoText(sheet,currRow))
        elif event == "noprocreate":
            greenRow += 1
            if greenRow == len(listGreen):
                break
            currRow = listGreen[greenRow][0].row
            window["text"].update(setInfoText(sheet,currRow))
        elif event == "cliplink":
            webview_process = multiprocessing.Process(target=launchLink,args=(sg.clipboard_get(),))
            webview_process.start()

    with open('greens.txt',"w") as f:
        f.writelines([str(x[1].row)+"\n" for x in listProcreate])
    
if __name__ == "__main__":
    main()

